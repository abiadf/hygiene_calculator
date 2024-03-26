'''Module that manages the Excel sheet: creating, filling, saving...'''

import openpyxl
import os

from config_info_obtainer import Constants
from logging_maker import logger
# from run_tempKPI_derivative import temp_abs_extrema, solution_type
from multi_file_maker import temp_abs_extrema, solution_type
# import phase_identifier_results as rpi
# from variables import Variables


class ExcelSheetMaker:
    '''Class that makes the excel sheet'''

    excel_extension = '.xlsx'

    def __init__(self, workbook_name, resulting_phases, filename):
        self.workbook_name : str  = workbook_name
        self.open_workbook        = [] #openpyxl.Workbook() #create workbook
        self.active_worksheet     = [] #self.open_workbook.active #select active worksheet

        self.filename      : str  = filename
        self.header_row_number:int= 1 # smallest value Excel rows can have
        self.header_exists : int  = 0
        self.loaded_file   : int  = None

        self.post_milk_flush_time = resulting_phases.post_milk_flush_time #this is from main
        self.post_milk_flush_idx  = resulting_phases.post_milk_flush_idx
        self.prerinse_time        = resulting_phases.prerinse_time
        self.prerinse_idx         = resulting_phases.prerinse_idx
        self.low_C_zone_KPIs     = resulting_phases.low_C_zone_KPIs
        self.hot_rinse_time       = resulting_phases.hot_rinse_time
        self.hot_rinse_idx        = resulting_phases.hot_rinse_idx
        self.post_rinse_time      = resulting_phases.post_rinse_time
        self.post_rinse_idx       = resulting_phases.post_rinse_idx
        self.rinse_KPIs        = resulting_phases.rinse_KPIs

        self.header_values = ['File name',
                              'Day of measurement',
                              'Start of post-milk flush [RT]',
                              'Start of pre-rinse [RT]',
                              'Start of hot rinse [RT]',
                              'Time for max T [RT]',
                              'Start of post-rinse [RT]',
                              'Max T [C]',
                              f"Avg. T of {Constants.time_interval}s interval with highest T [C]",
                              f"Duration for which T>{Constants.T_crit}C [s]",
                              'Avg. C for hot rinse (with water) [mS/cm]', 
                              'Avg. C for hot rinse (no water) [mS/cm]',
                              'Avg. C for hot rinse (no water) [%]',
                              'Solution type']

        self.row_values = [self.filename,
                           self.post_milk_flush_time.strftime('%Y-%m-%d'),
                           self.post_milk_flush_time.time(),
                           self.prerinse_time.time(),
                           self.hot_rinse_time.time(),
                           Variables.T_max_time.time(),
                           self.post_rinse_time.time(),
                           Variables.T_max,
                           temp_abs_extrema['T of max time interval [C]'],
                           temp_abs_extrema['Duration for which T > T_crit [s]'],
                           self.rinse_KPIs['C_avg hot rinse [mS/cm]'],
                           self.rinse_KPIs['C_avg hot rinse, no water [mS/cm]'],
                           self.rinse_KPIs['C_avg hot rinse, no water [%]'],
                           solution_type,]


    def find_existing_excel_files(self):
        '''Find Excel files in directory'''

        os.chdir(Constants.output_location)
        list_of_files_in_dir = os.listdir()

        excel_files_in_dir = []
        for file in list_of_files_in_dir:
            if file.endswith(self.excel_extension):
                excel_files_in_dir.append(file)

        logger.info(f"Excel files: {excel_files_in_dir}")
        return excel_files_in_dir


    def remove_existing_excel_files(self, excel_files_in_dir):
        '''Remove existing Excel files in directory'''

        for file in excel_files_in_dir:
                os.remove(file)
                logger.info(f"Removed file: {file}")

        return None


    def create_new_excel_file(self):
        '''Create new Excel file'''

        self.open_workbook   = openpyxl.Workbook() # create workbook
        self.active_worksheet= self.open_workbook.active # select active worksheet
        logger.info("Created new Excel file")
        return None


    def load_existing_excel_file(self, excel_files_in_dir):
        '''Load existing Excel file in directory'''

        if excel_files_in_dir:
            excel_file           = excel_files_in_dir[0] #take the first file
            self.open_workbook   = openpyxl.load_workbook(excel_file)
            self.active_worksheet= self.open_workbook.active #select active worksheet
            logger.info(f"Loaded Excel file '{excel_file}'")
            self.loaded_file = 1
        else:
            self.create_new_excel_file()
            logger.info(f"Did not find Excel file, so created {self.open_workbook}")
            self.loaded_file = 0

        return excel_file


    def check_if_header_row_filled(self):
        '''Looking through the rows of the Excel sheet to see which row is available to place results in
           Judging whether a row is filled based on the first cell in it (so column 1 or A)'''

        column_num = 1
        if (self.active_worksheet.cell(row = self.header_row_number, column = column_num).value ) != None:
            logger.info(f"Header exists on row {self.header_row_number}")
            self.header_exists = 1
        else:
            self.header_exists = 0

        return None


    def fill_header(self):
        '''Fill header row with values'''

        if self.header_exists == 0:
            for col, value in enumerate(self.header_values, start = 1):
                self.active_worksheet.cell(row = self.header_row_number, column = col, value = value)
            logger.info(f"Created header in row {self.header_row_number}!")
        else:
            logger.info(f"Header exists in row {self.header_row_number}, will not fill it!")

        return None


    def find_available_row(self, excel_file):
        '''Looping through the rows of the Excel sheet to see which row is available to place results in
           Judging whether a row is filled based on the first cell in it (so column 1 or A)'''

        print(excel_file)
        row_number = 1 # smallest row number in Excel
        counter = 0
        while True:
            if (self.active_worksheet.cell(row = row_number, column = 1).value ) == None:
                logger.info(f"Row {row_number} is empty")
                return row_number
            else:
                logger.info(f"Row {row_number} is NOT empty")
                row_number += 1
                counter += 1

            if counter == 200:
                logger.critical(f"There are more than {counter} entries in the Excel sheet. Shutting down")
                break


    def fill_row_with_values(self, free_row_number):
        '''Fill row with values'''

        for i, val in enumerate(self.row_values, start = 1):
            self.active_worksheet.cell(row = free_row_number, column = i, value = val)
            print(i, val)

        logger.info(f"Filled row {free_row_number}!")
        return None


    def save_workbook(self, excel_file):
        '''Save workbook to file. Saving depends on whether file was loaded or created'''

        if self.loaded_file == 0:
            self.open_workbook.save(self.workbook_name)
            logger.info(f"Saved a new workbook '{self.workbook_name}'")
        elif self.loaded_file == 1:
            self.open_workbook.save(excel_file)
            logger.info(f"Saved a loaded workbook '{excel_file}'")


    def close_workbook(self):
        '''Close workbook'''
        self.open_workbook.close()


    def make_excel_workbook(self, excel_file):
        '''Function that creates and saves the Excel workbook
        INPUT: header+cell values
        OUTPUT: None, creates and saves the excel file'''

        self.save_workbook(excel_file)
        self.close_workbook()
        return None


# ==============================================================================

    # def setup_excel_values(self):
    #     '''Function that makes the values to insert into the Excel workbook
    #     INPUT: workbook name
    #     OUTPUT: workbook values, to create a workbook out of'''

    #     workbook_values = [('A1', 'File name',                                   'A2', Constants.filename),
    #                        ('B1', 'Day of measurement',                          'B2', self.post_milk_flush_time.strftime('%Y-%m-%d')),
    #                        ('C1', 'Start of post-milk flush [RT]',               'C2', self.post_milk_flush_time.time()),
    #                        ('D1', 'Start of pre-rinse [RT]',                     'D2', self.prerinse_time.time()),
    #                        ('E1', 'Start of hot rinse [RT]',                     'E2', self.hot_rinse_time.time()),
    #                        ('F1', 'Time for max T [RT]',                         'F2', Variables.T_max_time.time()),
    #                        ('G1', 'Start of post-rinse [RT]',                 'G2', self.post_rinse_time.time()),
    #                        ('H1', 'Max T [C]',                                   'H2', Variables.T_max),
    #                        ('I1', f"Avg. T of {Constants.time_interval}s interval with highest T [C]",\
    #                                                                              'I2', temp_abs_extrema['T of max time interval [C]']),
    #                        ('J1', f"Duration for which T>{Constants.T_crit}C [s]",'J2', temp_abs_extrema['Duration for which T > T_crit [s]']),
    #                        ('K1', 'Avg. C for hot rinse (with water) [mS/cm]',   'K2', self.rinse_KPIs['C_avg hot rinse [mS/cm]']),
    #                        ('L1', 'Avg. C for hot rinse (no water) [mS/cm]',     'L2', self.rinse_KPIs['C_avg hot rinse, no water [mS/cm]']),
    #                        ('M1', 'Avg. C for hot rinse (no water) [%]',         'M2', self.rinse_KPIs['C_avg hot rinse, no water [%]']),
    #                        ('N1', 'Solution type',                               'N2', solution_type),]


    #     self.workbook_values = workbook_values
    #     return self.workbook_values


    # def make_excel_workbook(self):
    #     '''Function that creates and saves the Excel workbook
    #     INPUT: header+cell values
    #     OUTPUT: None, creates and saves the excel file'''

    #     open_workbook    = openpyxl.Workbook() #create workbook
    #     active_worksheet = open_workbook.active #select active worksheet

    #     # Fill Excel header with values
    #     header_location= 0
    #     header_name    = 1
    #     value_location = 2
    #     value_number   = 3
    #     for i, _ in enumerate(self.workbook_values):
    #         active_worksheet[self.workbook_values[i][header_location]] = self.workbook_values[i][header_name]
    #         active_worksheet[self.workbook_values[i][value_location]]  = self.workbook_values[i][value_number]

    #     # Auto-adjust column width
    #     number_of_columns  = len(self.workbook_values)
    #     alphabet_characters= string.ascii_uppercase
    #     for i in alphabet_characters[0 : number_of_columns]:
    #         active_worksheet.column_dimensions[i].auto_size = True

    #     open_workbook.save(self.workbook_name) # Save workbook to file

    #     logger.info(f"Created new Excel file '{self.workbook_name}' in {os.getcwd()}")
    #     logger.info(self.workbook_values)

    #     return None



'''Quick plots'''

# df_out = df_smooth
# # df_out.iloc[:, 1:4].plot()
# df_out.plot(x = df_out.columns[0], y=df_out.columns[1:4])
# plt.show()

# # df_diff.iloc[:, 1:4].plot()
# df_diff.plot(x = df_diff.columns[0], y=df_diff.columns[1:4])
# plt.show()

# # df_diff2.iloc[:, 1:4].plot()
# df_diff2.plot(x = df_diff2.columns[0], y=df_diff2.columns[1:4])
# plt.show()

